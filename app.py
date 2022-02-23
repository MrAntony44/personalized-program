from flask import Flask, render_template
from flask import request
app = Flask(__name__)
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import asyncio

aClasses = ["Α1", "Α2", "Α3", "Α4", "Α5"]
bClasses = ["Β1", "Β2", "Β3", "Β4", "Β5"]
cClasses = ["Γ1", "Γ2", "Γ3", "Γ4", "Γ5"]
sClasses = ["Γθετ1", "Γθετ2", "Γοικ1", "Γοικ2", "Γυγείας", "Γανθρ", "Βθετ1", "Βθετ2", "Βθετ3", "Βθετ4", "Βανθρ1", "Βανθρ2"]

classesAll = [
    {"c":"Α1","n":0},
    {"c":"Α2","n":1},
    {"c":"Α3","n":2},
    {"c":"Α4","n":3},
    {"c":"Α5","n":4},
    {"c":"Β1","n":5},
    {"c":"Β2","n":6},
    {"c":"Β3","n":7},
    {"c":"Β4","n":8},
    {"c":"Β5","n":9},
    {"c":"Γ1","n":10},
    {"c":"Γ2","n":11},
    {"c":"Γ3","n":12},
    {"c":"Γ4","n":13},
    {"c":"Γ5","n":14},
    {"c":"Γθετ1","n":15},
    {"c":"Γθετ2","n":16},
    {"c":"Γοικ1","n":17},
    {"c":"Γοικ2","n":18},
    {"c":"Γυγείας","n":19},
    {"c":"Γανθρ","n":20},
    {"c":"Βθετ1","n":21},
    {"c":"Βθετ2","n":22},
    {"c":"Βθετ3","n":23},
    {"c":"Βθετ4","n":24},
    {"c":"Βανθρ1","n":25},
    {"c":"Βανθρ2","n":26}
    
]

classes = aClasses + bClasses + cClasses + sClasses

wb = load_workbook("test.xlsx", data_only=True)
ws1 = wb["Εκπαιδευτικοί"]

ls0 = []
ls1 = []
ls2 = []
ls3 = []
ls4 = []

lsff = []

lastRow = 73
lastColumn = 36

# 2-8 its monday
# 9-15 its tuesday
# 16-22 its wednesday
# 23-29 its thursday
# 30-36 its friday

def mainFunction(si):
    ls0 = []
    ls1 = []
    ls2 = []
    ls3 = []
    ls4 = []
    count = 0

    for i in range(1, 80): # Row indexing
        for o in range(1, 38): # Column indexing
            if(ws1.cell(row=i, column=o).value == si):
                lC = i-1
                double = "0";
                if testMerge(ws1.cell(row=lC, column=o), ws1):
                    double = "1"                
                hour = ws1.cell(row=3, column=o).value
                lesson = ws1.cell(row=lC, column=o).value
                if o >= 2 and o<=8:
                    day = 0
                elif o >=9 and o<=15:
                    day = 1
                elif o >=16 and o<=22:
                    day = 2
                elif o >=23 and o<=29:
                    day = 3
                elif o >=30 and o<=36:
                    day = 4
                obj = {
                    "lesson": lesson,
                    "day": day,
                    "hour": hour,
                    "double": double
                }
                if day==0:
                    ls0.append(obj)
                elif day==1:
                    ls1.append(obj)
                elif day==2:
                    ls2.append(obj)
                elif day==3:
                    ls3.append(obj)
                elif day==4:
                    ls4.append(obj)
                count +=1
    lsF = {
        si: {
            "0": ls0,
            "1": ls1,
            "2": ls2,
            "3": ls3,
            "4": ls4
        }
    }
    lsff.append(lsF)

def testMerge(cell, sheet):
    for mergedCell in sheet.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return True
    return False

@app.route('/')
def index():
    return render_template("index.html")

def generateProgramJson():
    for i in classes:
        mainFunction(i)
    with open("data.json", 'w', encoding="utf8") as json_file:
        json.dump(lsff, json_file, indent=4, ensure_ascii=False, separators=(',',': '))

@app.route('/update')
def update():
  print ('updated json')
  generateProgramJson()
  return 'success'

@app.route('/find')
async def find():
    uClass = request.args.get('class')
    special = request.args.get('special')
    classId = 999
    specialId = 999
    specialExist = True
    if special == "none":
        specialExist = False
    print(uClass)
    for i in classesAll:
        if i["c"] == uClass:
            classId = i["n"]
    if specialExist:
        for i in classesAll:
            if i["c"] == special:
                specialId = i["n"]
    print(classId)
    print(specialId)
    print(specialExist)
    if classId == 999 or (specialId == 999 and specialExist):
        print("ERROR ON CLASSID AND SPECIALID")
        return
    a_file = open("data.json", "r", encoding='utf8')
    json_object = json.load(a_file)
    nClassProg = json_object[classId][uClass]
    if specialExist:
        sClassProg = json_object[specialId][special]
    else:
        sClassProg = []
    totalProg = {
        "normal": nClassProg,
        "special": sClassProg
    }
    a_file.close()
    print("found")
    return render_template('display.html', data=json.dumps(totalProg, ensure_ascii=False), args=json.dumps(request.args)) 

if __name__ == '__main__':
  app.run(debug=True)