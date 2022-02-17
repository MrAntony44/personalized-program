from flask import Flask, render_template
from flask import request
app = Flask(__name__)
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import asyncio

aClasses = ["Α1", "Α2", "Α3", "Α4", "Α5"]
bClasses = ["Β1", "Β2", "Β3", "Β4", "Β5"]
cClasses = ["Γ1", "Γ2", "Γ3", "Γ4", "Γ5"]
sClasses = ["Γθετ1", "Γθετ2", "Γοικ1", "Γοικ2", "Γυγείας", "Γανθρ", "Βθετ1", "Βθετ2", "Βθετ3", "Βθετ4", "Βανθρ1", "Βανθρ2"]

classes = aClasses + bClasses + cClasses + sClasses

wb = load_workbook("test.xlsx", data_only=True)
ws1 = wb["Εκπαιδευτικοί"]

ls0 = []
ls1 = []
ls2 = []
ls3 = []
ls4 = []

lastRow = 73
lastColumn = 36

# 2-8 its monday
# 9-15 its tuesday
# 16-22 its wednesday
# 23-29 its thursday
# 30-36 its friday

def mainFunction(si):
    ls0 = []
    ls1 = []
    ls2 = []
    ls3 = []
    ls4 = []
    count = 0

    for i in range(1, 73): # Row indexing
        for o in range(1, 37): # Column indexing
            if(ws1.cell(row=i, column=o).value == si):
                lC = i-1
                hour = ws1.cell(row=3, column=o).value
                lesson = ws1.cell(row=lC, column=o).value
                if o >= 2 and o<=8:
                    day = 0
                elif o >=9 and o<=15:
                    day = 1
                elif o >=16 and o<=22:
                    day = 2
                elif o >=23 and o<=29:
                    day = 3
                elif o >=30 and o<=36:
                    day = 4
                obj = {
                    "lesson": lesson,
                    "day": day,
                    "hour": hour
                }
                if day==0:
                    ls0.append(obj)
                elif day==1:
                    ls1.append(obj)
                elif day==2:
                    ls2.append(obj)
                elif day==3:
                    ls3.append(obj)
                elif day==4:
                    ls4.append(obj)
                count +=1
    lsF = {
        si: {
            "Monday": ls0,
            "Tuesday": ls1,
            "Wednesday": ls2,
            "Thursday": ls3,
            "Friday": ls4
        }
    }
    print(lsF)
    file_object = open('data.json', 'a', encoding='utf8')
    file_object.write(json.dumps(lsF, ensure_ascii=False))

    print("done")

def generateProgramJson():
    file_object = open('data.json', 'w', encoding='utf8')
    file_object.write(json.dumps('', ensure_ascii=False))
    for i in classes:
        mainFunction(i)


@app.route('/')
def index():
    # mainFunction()
    return render_template("index.html")
  

@app.route('/update')
async def update():
  print ('updated json')
  generateProgramJson()
  return 'update done'

@app.route('/find')
async def find():
    uClass = request.args.get('class')
    special = request.args.get('special')

    a_file = open("data.json", "r", encoding='utf8')
    # json_object = json.load(a_file)
    a_file.close()
    # print(json_object)
    return 'update done'


if __name__ == '__main__':
  app.run(debug=True)

